import os, sys, time, csv, itertools
from collections import defaultdict
import numpy as np

import openpyxl

reads_threshold = 10.0 
COUNT = 0
minor_alleles_and_hybrids = defaultdict(list)
true_allelelist = {}
prev_dataset = ""


#----------- OBTAIN HYBRIDS AND ITS PARENTS FROM A DIRECTORY -----------#

def create_mito_features_dataset_from_dir(directory, pcr_kit, train=True):
    '''
    This function creates feature datasets for hybrid noise, non-hybrid noise 
    (other noise) and minor/hybrid allele combinations. From the dir input it 
    fetches all the subdirs conatining output from FDSTools. It reads the 
    contents of it and compares it to the true alleles in the dataset, to find 
    possible alleles from a third contributor that can also be formed as hybrid.
    The output is written to CSV files over the complete dataset given and as 
    fragments, with the corresponding labels file with the hybrid ratio.
    '''
    global minor_alleles_and_hybrids
    columns = ['hybrid_reads', 'highest_reads_par', 'lowest_reads_par', \
              'ratio_parents', 'len_hybrid', 'len_longest_par', \
              'len_shortest_par', 'begin_pos_overlap', 'end_pos_overlap', \
              'len_overlap', 'longest_c_stretch', 'len_overlap_parA', \
              'len_overlap_parB', 'hybrid_CG_counts', 'hybrid_AT_counts', \
              'overlap_CG_counts', 'overlap_AT_counts', 'hybrid_A_counts', \
              'hybrid_C_counts', 'hybrid_G_counts', 'hybrid_T_counts', \
              'overlap_A_counts', 'overlap_C_counts', 'overlap_G_counts', \
              'overlap_T_counts', 'MT_hybrid', 'MT_overlap', 'hybrid_mol_weight']

    all_features = []
    all_labels = []
    minor_features =[]
    minor_labels =[]
    start_time = time.time()
    dirs = []

    kit = directory.split('/')[-2]
    hybrid_dd = defaultdict(list)
    hybrid_label_dd = defaultdict(list)

    # not used in thesis (other noise!)
    non_hybrid_dd = defaultdict(list)
    non_hybrid_label_dd = defaultdict(list)

    minor_dd = defaultdict(list)
    minor_label_dd = defaultdict(list)

    for (root,_,_) in os.walk(directory): 
        if "results_fdstools_" in root:
            dirs.append(root)

    for dir in dirs:
        # this statments excludes specific data sets and every thing that does
        # not have the dataset_name in it
        pcr_kit_year = dir.replace(directory,"").split('/')[0]
        if "_mito-mini_low" in dir or 'Saliva' in dir or not pcr_kit in pcr_kit_year:
            continue
        if "2023" in pcr_kit_year:
            continue

        hybrid_features, non_hybrid_features = obtain_all_hybrids_from_dir(dir, pcr_kit_year, train)

        # creates a label and feature dictionary from hybrids 
        for key, values in hybrid_features.items():
            for value in values:
                hybrid_dd[key].append(value[0])
                hybrid_label_dd[key].append(value[1])
                
        # creates a label and feature dictionary from other noise (not for thesis)
        for key, values in non_hybrid_features.items():
            for value in values:
                non_hybrid_dd[key].append(value[0])
                non_hybrid_label_dd[key].append(value[1])
        

        # creates a label and feature dictionary from the hybrid/minor combis
        for key, values in minor_alleles_and_hybrids.items():
            for value in values:
                minor_dd[key].append(value[0])
                minor_label_dd[key].append(value[1])

        minor_alleles_and_hybrids = defaultdict(list)

        
        time_elapsed = time.time()-start_time
        print('Obtained features of {:s} in {:.0f}m {:.0f}s'.format(dir, time_elapsed // 60, time_elapsed % 60))

    # Comment out to inspect linearity in data
    # pp.histogram_ratios_per_class(hybrid_label_dd, non_hybrid_label_dd)
    # pp.plot_ratio_to_reads(hybrid_dd,hybrid_label_dd, non_hybrid_dd, \
    # non_hybrid_label_dd)


    kit = 'Features_data_' + kit + "_" + pcr_kit + '_new'

    for marker, hybrid_features in hybrid_dd.items():
        all_features.extend(hybrid_features)

        minor_features.extend(minor_dd[marker])
        minor_labels.extend(minor_label_dd[marker])

        hybrid_labels = hybrid_label_dd[marker]        
        all_labels.extend(hybrid_labels)
        
        num_of_hybrids = len(hybrid_labels)
        print(f'{marker} has {num_of_hybrids} hybrids')

        # write marker features to a csv file 
        csv_marker_features_name = f'{kit}/features_{marker}.csv'
        write_to_file(csv_marker_features_name,hybrid_features,columns)

        # write marker labels to a csv file
        csv_marker_labels_name = f'{kit}/labels_{marker}.csv'
        labels = list(map(lambda el:[el], hybrid_labels))
        write_to_file(csv_marker_labels_name,labels,["labels"])
        
    
    # pkl_file_name = f'{kit}/features_minor.pkl' 
    # df_features.to_pickle(pkl_file_name)
    csv_minor_features_name = f'{kit}/features_minor.csv'
    write_to_file(csv_minor_features_name,minor_features,columns)


    # labels_file = f'{kit}/labels_minor.pkl' 
    # df_labels.to_pickle(labels_file)
    csv_minor_labels_name = f'{kit}/labels_minor.csv'
    minor_labels = list(map(lambda el:[el], minor_labels))
    write_to_file(csv_minor_labels_name,minor_labels,["labels"])
    
    # pkl_file_name = f'{kit}/features_all.pkl' 
    # df_features.to_pickle(pkl_file_name)
    csv_features_all_name = f'{kit}/features_all.csv'
    write_to_file(csv_features_all_name,all_features,columns)

    # labels_file = f'{kit}/labels_all.pkl' 
    # df_labels.to_pickle(labels_file)
    csv_labels_all_name = f'{kit}/labels_all.csv'
    all_labels = list(map(lambda el:[el], all_labels))
    write_to_file(csv_labels_all_name,all_labels,["labels"])

    return csv_features_all_name, csv_labels_all_name    

def obtain_all_hybrids_from_dir(dir, pcr_kit_year, train=True):
    '''
    This function takes one FDSTools output direcotry and then obtains all the 
    potential hybrids in each file and outputs them as a default dictionaries 
    for the hybrid noise, non-hybrid noise and a list of hybird parents. This 
    last list was used for testing and might be obsolete.
    '''
    dir_name = dir.split('_')
    hybrid_feature_dict = defaultdict(list)
    non_hybrid_feature_dict = defaultdict(list)
    # creates dictionaries for hybrids, hybrid_features and non-hybrid features
    for file in os.listdir(dir):
        if file.endswith(".txt"):
            if train:
                # Exclude the two files used for validation in the thesis. These  
                # are a 2-pers mix and 3-pers mix from 2 unseen individuals.
                if ("2023" in dir_name[0]):
                    if "Mito-mini" in pcr_kit_year and (file == 'T0216-table.txt'\
                                                     or file == 'T0320-table.txt'):
                        continue
                    elif "Easymito" in pcr_kit_year and (file == 'E0122-table.txt'\
                                                     or file == 'E0155-table.txt'):
                        continue
            # if "2018" in dir_name[0]:
            #     continue
            
            file_path = os.path.join(dir, file)
                    
            hybrid_features, non_hybrid_features = obtain_all_hybrids_from_file(file_path, pcr_kit_year, train)

            hybrid_feature_dict = merge_defaultdicts(hybrid_feature_dict, hybrid_features)
            non_hybrid_feature_dict = merge_defaultdicts(non_hybrid_feature_dict, non_hybrid_features)

    return hybrid_feature_dict, non_hybrid_feature_dict, 

def merge_defaultdicts(d,d1):
    '''
    https://stackoverflow.com/questions/31216001/how-to-concatenate-or-combine-two-defaultdicts-of-defaultdicts-that-have-overlap
    '''
    for k,v in d1.items():
        if (k in d):
            d[k].extend(v)
        else:
            d[k] = d1[k]
    return d

def obtain_all_hybrids_from_file(filename, pcr_kit_year, train=True, f=''):
    marker_hybrid_dict = {}
    global minor_alleles_and_hybrids
    hybrid_features_dict = defaultdict(list)
    non_hybrid_features_dict = defaultdict(list)
    filepath_list = filename.split('/')
    if not train:
        dataset = ''
    barcode_table = filepath_list[-1].replace('-table.txt', "")

    marker_sequences_dict = create_sequence_dict_from_file(filename)
    # marker_sequences_dict  = create_sequence_dict_from_file(filename)

    if marker_sequences_dict == {}:
        return {},{}

    for marker, values in marker_sequences_dict.items():
        ''' #comment this line if you want to pretty print a file
        f = open(f"{barcode_table}", "at")
        f.write('*'*40 + '\n' + f'{marker}' +   '\n')
        f = open(f"{filename}-{marker}", "at")
        #'''

        # create features for the  hybrids with parents and non-hybrids with parents
        align_hybrid_parents, align_non_hybrid_parents = get_hybrid_parents_combis_and_overlaps(values, train, f)
        if align_hybrid_parents != []:

            if marker not in hybrid_features_dict:
                marker_hybrid_dict[marker] = []
                hybrid_features_dict[marker] = []
            
            if train:
                minor_alleles = get_minor_allele_hybrid_combi(align_hybrid_parents, marker, barcode_table, pcr_kit_year)
                if minor_alleles != []:
                    if marker not in minor_alleles_and_hybrids:
                        minor_alleles_and_hybrids[marker] = []
                    minor_alleles_and_hybrids[marker].extend(get_features_and_labels_from_hybrid(minor_alleles, marker))
                    for i in minor_alleles:
                        align_hybrid_parents.remove(i)

            # marker_hybrid_dict was needed for testing, not feature creation
            # it contains alle hybrid parents combinations
            marker_hybrid_dict[marker].extend(align_hybrid_parents) 
            features = get_features_and_labels_from_hybrid(align_hybrid_parents, marker)
            hybrid_features_dict[marker].extend(features)

        # not used in thesis:
        if align_non_hybrid_parents != []:
            if marker not in non_hybrid_features_dict:
                non_hybrid_features_dict[marker] = []
            features = get_features_and_labels_from_hybrid(align_non_hybrid_parents, marker)
            non_hybrid_features_dict[marker].extend(features)

            
    return hybrid_features_dict, non_hybrid_features_dict

def create_sequence_dict_from_file(file_path):
    tssv = False
    if 'tssv' in file_path:
        tssv = True
    
    markers_sequences_dict = {}
    allele_dict = {}
    prev_marker = "Frag01_mt16009-16129"
    with open(file_path) as seqfile:
        next(seqfile)
        for line in seqfile:
            line = line.rstrip("\r\n").split("\t")
            if line == ['']: # needed for newer version of FDSTools
                markers_sequences_dict[prev_marker] = allele_dict
                allele_dict = {}
                return markers_sequences_dict
                
            marker = line[0]
            if marker != prev_marker:
                markers_sequences_dict[prev_marker] = allele_dict
                allele_dict = {}
            
            if tssv:
                true_allele = ''
                sequence = line[1]
                reads = line[2]
            else:
                true_allele = line[1]
                reads = line[22]
                sequence = line[3]
            if sequence == "":
                continue
            if float(reads) < reads_threshold: #of <= ?
                continue
            seq_reads_dict = {"sequence":true_allele, "reads":reads}
            allele_dict[sequence] = seq_reads_dict

            prev_marker = marker

    return markers_sequences_dict
# create_sequence_dict_from_file
            

def get_hybrid_parents_combis_and_overlaps(sequence_reads_dict, train=True, f=''):
    global COUNT
    hybrid_parents_combis = []
    non_hybrid_parent_combis = []
    sequence_list = list(sequence_reads_dict.keys())

    # Consider each sequence  and check if it can be a hybrid
    for pos_hybrid in sequence_list:
    
        #creates a possible parent list without the possible hybrid
        parent_sets = []
        temp = sequence_list[:]
        temp.remove(pos_hybrid)

        # pick two parents from the left over sequences 
        # parent_sets = parent_sets + [(i,i) for i in temp] # used to get hybrids from one parent
        parent_sets = parent_sets + [i for i in itertools.combinations(temp,2)]

        for parent_set in parent_sets:
            hybrid_parents_read_list, hybrid_bool = get_hybrid_parents_list(pos_hybrid, parent_set, sequence_reads_dict, train) 
            if hybrid_parents_read_list == []:
                continue
            if hybrid_bool:
                COUNT += 1
                hybrid_parents_combis.append(hybrid_parents_read_list)
            else:
                non_hybrid_parent_combis.append(hybrid_parents_read_list)

    '''
    # comment the 3 ' above if you want to pretty print the hybrid/parents
    # and create its text file
    for i in hybrid_parents_combis:
        if isinstance(i, int):
            print("Length overlap = " + str(i))
            f.write("Length overlap = " + str(i))
            continue
        if isinstance(i, str):
            print("Allele hybrid: " + i)
            f.write("Allele hybrid: " + i)
            continue
        pp.pretty_print_alignments(i,f)
        print('\n')
        f.write('\n'))
    f.close()

    # '''
    return hybrid_parents_combis, non_hybrid_parent_combis
#get_hybrid_parents_combis_and_overlaps

def get_hybrid_parents_list(pos_hybrid, parent_set, sequence_reads_dict, train=True):
    '''
    Returns the hybrid aligned with the two parents, overlapping pieces, allele 
    names of hybrid and the parents and checks if it is hybrid noise or other 
    noise. Also checks the assumption for training where the parent reads must 
    be higher than hybrid reads.
    '''
    reads_hybrid = float(sequence_reads_dict.get(pos_hybrid)['reads'])
    reads_parA = float(sequence_reads_dict.get(parent_set[0])['reads'])
    reads_parB = float(sequence_reads_dict.get(parent_set[1])['reads'])

    if train:
        # the hybrid reads must be lower than both parent reads for the training set
        if reads_hybrid >= reads_parA or reads_hybrid >= reads_parB:
            return [], False
    else:
        # the hybrid and parent reads must at least be 30 reads
        th_reads = 30
        if reads_hybrid < th_reads or reads_parA < th_reads  or reads_hybrid < th_reads:
            return [], False
        if reads_hybrid >= reads_parA or reads_hybrid >= reads_parB:
            return [], False


    # align the hybrid with the parents
    align_hybrid_parents, overlap, overlap_A, overlap_B, hybrid_bool = align_3_sequences_prefix_suffix(pos_hybrid, parent_set)

    # add hybrid/parents combi and stop iterating with this hybrid and parent set
    if align_hybrid_parents != []:
        align_hybrid_parents_read_dict = {}
        if align_hybrid_parents[1] == parent_set[0]:
            align_hybrid_parents_read_dict = {align_hybrid_parents[0]: reads_hybrid,\
             align_hybrid_parents[1]: reads_parA, align_hybrid_parents[2]: reads_parB} 
        else:
            align_hybrid_parents_read_dict = {align_hybrid_parents[0]: reads_hybrid,\
             align_hybrid_parents[1]: reads_parB, align_hybrid_parents[2]: reads_parA} 
        
        name_hybrid = sequence_reads_dict.get(pos_hybrid)['sequence'].split('_')[0]
        name_parA = sequence_reads_dict.get(parent_set[0])['sequence'].split('_')[0]
        name_parB = sequence_reads_dict.get(parent_set[1])['sequence'].split('_')[0]
        return  [align_hybrid_parents_read_dict, overlap, overlap_A, overlap_B, \
                 name_hybrid, name_parA, name_parB], hybrid_bool

    return [], False

#--------------- ALIGNMENT OF SEQUENCES ---------------#

def align_3_sequences_prefix_suffix(hybrid,parent_set):
    '''
    This function will find the longest common prefix and suffix of parents with
    the hybrid in both ways and compares them with the hybrid to see if a hybrid
    might have formed or not. It returns the hybrid and parents in order; prefix
    parent, suffix parent and also returns the overlapping piece, the prefix
    and suffix and a bool if it is hybrid noise or other noise.
    '''

    parA, parB =  parent_set[0], parent_set[1]

    # get possible suffix and prefix from both parent sequences
    parA_prefix = longest_common_prefix_suffix([hybrid, parA])
    parB_prefix = longest_common_prefix_suffix([hybrid, parB])
    parA_suffix = longest_common_prefix_suffix([hybrid[::-1], parA[::-1]])
    parB_suffix = longest_common_prefix_suffix([hybrid[::-1], parB[::-1]])

    #dubbele code --> eruit slopen! --> ??? --> nog doen?
    if len(parA_prefix + parB_suffix) > len(hybrid):
        parB_suffix_rev = parB_suffix[::-1]
        overlap = find_overlapping_region(parA_prefix, parB_suffix_rev)
        parB_suffix_rev_no_overlap = parB_suffix_rev[len(overlap):]

        #check if hybrid and return hybrid bool as True for hybrid creation
        if parA_prefix + parB_suffix_rev_no_overlap == hybrid:
            return [hybrid,parA, parB], overlap, parA_prefix, parB_suffix[::-1], True
        else:
            return [hybrid,parA, parB], overlap, parA_prefix, parB_suffix[::-1], False

    if len(parB_prefix + parA_suffix) > len(hybrid):
        parA_suffix_rev = parA_suffix[::-1]
        overlap = find_overlapping_region(parB_prefix, parA_suffix_rev)
        if parB_prefix + parA_suffix_rev[len(overlap):] == hybrid:
            return [hybrid,parB,parA], overlap,  parB_prefix, parA_suffix[::-1], True
        else:
            return [hybrid,parB,parA], overlap,  parB_prefix, parA_suffix[::-1], False

    return [], 0,0,0, False
        

def find_overlapping_region(seq1, seq2):
    overlap = ''
    for nt in range(len(seq1)):
        if seq2.startswith(seq1[nt:]):
            overlap = seq1[nt:]
            break   
    # The overlap can not be the entire sequence
    if overlap == seq1 or overlap == seq2:
        return ''
    return overlap

def longest_common_prefix_suffix(two_sequences):
    two_sequences.sort()
    end = min(len(two_sequences[0]), len(two_sequences[1]))
    i = 0
    while i<end and two_sequences[0][i] == two_sequences[1][i]:
        i+=1
    prefix = two_sequences[0][0: i]
    return prefix
    

#-------------- TRUE ALLELES ---------------#

def read_mito_true_alleles(excel_file):
    '''
    This function reads in the excel file with the true alleles per sample for
    a given dataset. It returns all the true alleles per fragment per barcode.
    '''
    # Initiate dictionaries
    sample_content_dict = {}
    true_alleles_per_sample_dict = {}
    genotypes_dict = {}

    # Open sheets in the excel file with the true alleles.
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    sample_content = wb.worksheets[0]
    genotypes = wb.worksheets[1]

    # Fill dictionary with the individuals per sample
    for i, _ in enumerate(sample_content):  
        if i < 5:
            continue
        tag = sample_content.cell(i, 1).value
        samples = sample_content.cell(i, 2).value
        sample_content_dict[tag] = str(samples).split('+')

    # Create the marker list
    fragment_list = []
    for i, individual in enumerate(genotypes["A"]):
        if i == 0:
            continue
        fragment_list.append(individual.value)

    # Fill dictionary with alleles per individual
    for col in genotypes.iter_cols(min_row=1, min_col=2, values_only=True):
        alleles = list(col[1:])
        alleles_dict = dict(zip(fragment_list, alleles))
        genotypes_dict[col[0]] = alleles_dict

    # Fill dictionary with alleles per barcode
    for barcode, donors in sample_content_dict.items():
        if barcode not in true_alleles_per_sample_dict:
            true_alleles_per_sample_dict[barcode] = {}
        true_alleles_list = []
        for donor, true_alleles in genotypes_dict.items():  
            if str(donor) in donors:
                true_alleles_list.append(true_alleles)
        dd = defaultdict(list)
        for individual in true_alleles_list:
            for fragment, alleles in individual.items():
                dd[fragment].append(alleles)  
        true_alleles_per_sample_dict[barcode] = dd
    return true_alleles_per_sample_dict


def give_true_alleles_per_dataset(true_alleles_file):
    '''Create a nested dictionary containg all true allele sequences per 
    fragment per sample.'''
    true_alleles_sample_dict = {}

    # open the txt file with all true alleles for all fragments for all samples
    with open(true_alleles_file) as f:  
        next(f)
        for line in f:
            true_allele_sample = line.replace("\n", "").split('\t')
            sample = true_allele_sample[0]
            marker = true_allele_sample[1]
            allele = true_allele_sample[2]
            if sample not in true_alleles_sample_dict:
                true_alleles_sample_dict[sample] = {marker : []}
            if marker not in true_alleles_sample_dict[sample].keys():
                true_alleles_sample_dict[sample].update({marker:[]})
          
            true_alleles_sample_dict[sample][marker].append(allele)
    
    return true_alleles_sample_dict

def get_minor_allele_hybrid_combi(list_hybrid_parents, marker, tag, dataset):
    '''
    This function finds all minor alleles, that are a true allele from a 
    third contibutor as well. This function requires a txt file with the true 
    alleles of each sample in the training data set, that is aligned according
    to the allelelist txt files.
    '''
    minor_allele_hybrid = []
    global true_allelelist
    global prev_dataset

    # find all true alleles for each given dataset
    if true_allelelist == {} or prev_dataset != dataset:
        prev_dataset = dataset
        filename = f"True_alleles/allelelist_{dataset}.txt"
        true_allelelist = give_true_alleles_per_dataset(filename) 

    # this for loop compares al the hybrid alleles with the true alleles and 
    # finds all the hybrids that are actual alleles as well
    for hybrid_allele in list_hybrid_parents:
        if tag in true_allelelist:
            samples = true_allelelist[tag]
            true_alleles = samples[marker]
            true_hybrid_allele = list(hybrid_allele[0].keys())[0]
            if true_hybrid_allele in true_alleles:
                minor_allele_hybrid.append(hybrid_allele)
    return minor_allele_hybrid


#-------------- HELPER FUNCTIONS ---------------#

def write_to_file(csv_file_name, rows, columns):
    if not os.path.exists(csv_file_name.split('/')[0]):
        os.makedirs(csv_file_name.split('/')[0])


    with open(csv_file_name, "w") as f:
        writer = csv.writer(f)
        writer.writerow(columns)
        for row in rows:
            writer.writerow(row)




#--------------- FEATURE DATASET CREATION ---------------#

MW_dATP = 313.2
MW_dCTP = 289.2
MW_dGTP = 329.2
MW_dTTP = 304.2
MW_addition = 79.0

df_MT = {'AA': {'H': -9.1, 'S': -0.024}, 'AT': {'H': -8.6, 'S': -0.0239},\
        'TA': {'H': -6.0, 'S': -0.0169}, 'CA': {'H': -5.8, 'S': -0.0129}, \
        'GT': {'H': -6.5, 'S': -0.0173}, 'CT': {'H': -7.8, 'S': -0.0208}, \
        'GA': {'H': -5.6, 'S': -0.0135}, 'CG': {'H': -11.9, 'S': -0.0278}, \
        'GC': {'H': -11.1, 'S': -0.0267}, 'GG': {'H': -11.0, 'S': -0.0266}, \
        'TT': {'H': -9.1, 'S': -0.024}, 'CC': {'H': -11.0, 'S': -0.0266}, \
        'AG': {'H': -7.8, 'S': -0.0208}, 'AC': {'H': -6.5, 'S': -0.0173}, \
        'TC': {'H': -5.6, 'S': -0.0135}}

def calculate_molecular_weight(a_counts, c_counts, g_counts, t_counts):
    mol_weight = (a_counts*MW_dATP) + (t_counts*MW_dTTP) +(c_counts*MW_dCTP) + \
        (g_counts*MW_dGTP) + MW_addition
    return mol_weight

def marmor_doty_formula(a_counts, c_counts, g_counts, t_counts):
    # Marmor-Doty forula (1962) / Wallace rule
    return 2 * (a_counts + t_counts) + 4 *(c_counts+g_counts) - 7

def nearest_neighbour_formula(sequence):
    # Nearest Neigbours melting temperatuur len(sequence >= 14)
    sum_H = 0.0
    sum_S = 0.0
    for i in range(1,len(sequence)):
        pair = sequence[i-1] + sequence[i]
        if pair in df_MT.keys():
            sum_H += df_MT[pair]["H"]
            sum_S += df_MT[pair]["S"]

    # 0.0108 = constant for helix initation
    # 0.00199 = gas constant
    # 0.0000005 = oligonucleotide concentration in M
    return sum_H / (-0.0108 + sum_S + 0.00199 * np.log(0.0000005/4)) - 273.15 + 16.6*np.log10(0.05)


def find_longest_c_stretch(seq):
    maximum = count = 0
    for c in seq:
        if c == 'C':
            count += 1
        else:
            count = 0
        maximum = max(count,maximum)
    return maximum
 
#--------------- FEATURE DATASET CREATION ---------------#

def get_features_and_labels_from_hybrid(hybrid_parents_overlap_allele, marker):
    samples = []
    global pop_freq

    for i in hybrid_parents_overlap_allele:
        allele_freq = 0
        all_sequences = i[0]
        overlap = i[1]
        overlap_parA = i[2]
        overlap_parB = i[3]
        true_allele = i[4]

        sequences =  list(all_sequences.keys())
        reads =  list(all_sequences.values())


        hybrid_sequence = sequences[0]

        hybrid_reads = float(reads[0])
        parA_reads = float(reads[1])
        parB_reads = float(reads[2])

        float_reads = [parA_reads, parB_reads]
        parent_seqs = [sequences[1], sequences[2]]

        index_highest_reads = float_reads.index(max(float_reads))
        index_lowest_reads = float_reads.index(min(float_reads))

        parent_highest_reads = parent_seqs[index_highest_reads]
        parent_lowest_reads = parent_seqs[index_lowest_reads]

        highest_reads_par = float_reads[index_highest_reads]
        lowest_reads_par = float_reads[index_lowest_reads]
        
        ratio_parents = lowest_reads_par/highest_reads_par

        len_hybrid = len(hybrid_sequence)
        len_longest_par= max(len(sequences[1]), len(sequences[2]))
        len_shortest_par = min(len(sequences[1]), len(sequences[2]))

        len_overlap = int(len(overlap))
        len_overlap_parA = int(len(overlap_parA))
        len_overlap_parB = int(len(overlap_parB))
        marker_pos = int(marker.split('_')[1].replace("mt","").split('-')[0])
        begin_pos_overlap = marker_pos + (len_hybrid - len_overlap_parB)
        end_pos_overlap = marker_pos + len_overlap_parA

        hybrid_A_counts = hybrid_sequence.count('A')
        hybrid_C_counts = hybrid_sequence.count('C')
        hybrid_G_counts = hybrid_sequence.count('G')
        hybrid_T_counts = hybrid_sequence.count('T')

        overlap_A_counts = overlap.count('A')
        overlap_C_counts = overlap.count('C')
        overlap_G_counts = overlap.count('G')
        overlap_T_counts = overlap.count('T')

        longest_c_stretch = find_longest_c_stretch(hybrid_sequence)

        hybrid_CG_counts = hybrid_sequence.count('CG') 
        hybrid_AT_counts = hybrid_sequence.count('AT')
        hybrid_GC_counts = hybrid_sequence.count('GC') 
        hybrid_TA_counts = hybrid_CG_counts = hybrid_sequence.count('TA') 
        overlap_CG_counts = overlap.count('CG') 
        overlap_AT_counts = overlap.count('AT')
        overlap_GC_counts = overlap.count('GC') 
        overlap_TA_counts = overlap.count('TA')


        MT_hybrid = marmor_doty_formula(hybrid_A_counts,hybrid_C_counts,hybrid_G_counts,hybrid_T_counts) \
            if len_hybrid >= 14 else nearest_neighbour_formula(hybrid_sequence)
        MT_overlap = marmor_doty_formula(overlap_A_counts,overlap_C_counts,overlap_G_counts,overlap_T_counts) \
            if len_overlap >= 14 else nearest_neighbour_formula(overlap)

        hybrid_mol_weight = calculate_molecular_weight(hybrid_A_counts, hybrid_C_counts,hybrid_G_counts, hybrid_T_counts)
        
        # final feature list
        sample = [hybrid_reads, highest_reads_par, lowest_reads_par,\
            ratio_parents, len_hybrid, len_longest_par, len_shortest_par,\
            begin_pos_overlap, end_pos_overlap, len_overlap, longest_c_stretch,\
            len_overlap_parA, len_overlap_parB, hybrid_CG_counts,\
            hybrid_AT_counts, overlap_CG_counts, overlap_AT_counts,\
            hybrid_A_counts, hybrid_C_counts, hybrid_G_counts, hybrid_T_counts,\
            overlap_A_counts, overlap_C_counts, overlap_G_counts,\
            overlap_T_counts, MT_hybrid, MT_overlap, hybrid_mol_weight]

        # label creation
        label = hybrid_reads/(parA_reads + parB_reads)
        sample = [sample,label, parent_highest_reads, parent_lowest_reads, hybrid_sequence]
        samples.append(sample)

    return samples


#--------------- MAIN FUNCTIONS ---------------#

def main():
    dir = sys.argv[1]
    dataset_name = sys.argv[2]
    create_mito_features_dataset_from_dir(dir, dataset_name)

    global COUNT
    print(f"Total number of hybrids: {COUNT}")


if __name__ == "__main__":
    main()

